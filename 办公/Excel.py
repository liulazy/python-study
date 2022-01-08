import openpyxl
filename = r'C:\Users\奥子gei\Downloads\刘金澳-2311162708.xlsx'         #Excel路径
wb = openpyxl.load_workbook(filename)

#ws = wb.get_sheet_by_name('Sheet1')老方法
ws = wb['Sheet1']
maxrow=ws.max_row
j = 1
for i in range(2,maxrow+1):
    value1 = ws.cell(i, j).value
    print(value1)


#value1 = ws.cell(i,j).value       #读值
#print(value1)
#ws.cell(i,j).value = 10            #赋值
#wb.save(filename+'22.xlsx')        #创建一个新的Excel